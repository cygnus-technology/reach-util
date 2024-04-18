import argparse
import builtins
from openpyxl import load_workbook
import re

import c_code_generator_utils as ccgu

SCRIPT_MAJOR_VERSION = 0
SCRIPT_MINOR_VERSION = 0
SCRIPT_PATCH_VERSION = 1

INDENT_SIZE = 4
INDENT = f"{' ' * INDENT_SIZE}"


def gen_header(filename, brief):
    return r'''/********************************************************************************************
 *    _ ____  ___             _         _     ___              _                        _
 *   (_)__ / | _ \_ _ ___  __| |_  _ __| |_  |   \ _____ _____| |___ _ __ _ __  ___ _ _| |_
 *   | ||_ \ |  _/ '_/ _ \/ _` | || / _|  _| | |) / -_) V / -_) / _ \ '_ \ '  \/ -_) ' \  _|
 *   |_|___/ |_| |_| \___/\__,_|\_,_\__|\__| |___/\___|\_/\___|_\___/ .__/_|_|_\___|_||_\__|
 *                                                                  |_|
 *                           -----------------------------------
 *                          Copyright i3 Product Development 2024
 *
 * MIT License
 *
 * Permission is hereby granted, free of charge, to any person obtaining a copy
 * of this software and associated documentation files (the "Software"), to deal
 * in the Software without restriction, including without limitation the rights
 * to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
 * copies of the Software, and to permit persons to whom the Software is
 * furnished to do so, subject to the following conditions:
 *
 * The above copyright notice and this permission notice shall be included in all
 * copies or substantial portions of the Software.
 *
 * THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
 * IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
 * FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
 * AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
 * LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
 * OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
 * SOFTWARE.
 *
 * @file      ''' + f"{filename}" + r'''
 * @brief     ''' + f"{brief}" + r'''
 * @copyright 2023-2024 i3 Product Development. All Rights Reserved.
 *
 * Original Author: Chuck Peplinski
 * Script Author: Joseph Peplinski
 *
 * Generated with version ''' + \
        f"{SCRIPT_MAJOR_VERSION}.{SCRIPT_MINOR_VERSION}.{SCRIPT_PATCH_VERSION}" + ''' of c_code_generator.py
 *
 ********************************************************************************************/'''


def make_c_compatible(name: str, upper=False):
    if upper:
        name = name.upper()
    else:
        name = name.lower()
    return re.sub(r'\W+|^(?=\d)', '_', name).removesuffix("_")


def gen_enum(enums, values, name, indent=0, transform_enum_names=False):
    if transform_enum_names:
        for i in range(len(enums)):
            raw_enum = f"{name}_{enums[i]}"
            enums[i] = make_c_compatible(raw_enum, upper=True)
    lines = [f"{INDENT * indent}typedef enum {{"]
    max_len = len(max(enums, key=len))
    for enum, val in zip(enums, values):
        if val is not None:
            lines.append(f"{INDENT * (indent + 1)}{enum.ljust(max_len)} = {val},")
        else:
            lines.append(f"{INDENT * (indent + 1)}{enum},")
    lines.append(f"{INDENT * indent}}} {make_c_compatible(name)}_t;")
    separator = "\n"
    return separator.join(lines)


class CrDeviceInfo:
    def __init__(self, data):
        if data["Device Name"]:
            self.device_name = data["Device Name"]
        if data["Manufacturer"]:
            self.manufacturer = data["Manufacturer"]
        if data["Device Description"]:
            self.device_description = data["Device Description"]
        if data["CLI Supported"] == "Yes":
            self.cli = True
        elif data["CLI Supported"] == "No":
            self.cli = False
        else:
            raise ValueError(f"Invalid 'CLI Supported' value {data['CLI Supported']}")
        if data["Time Supported"] == "Yes":
            self.time = True
        elif data["Time Supported"] == "No":
            self.time = False
        else:
            raise ValueError(f"Invalid 'Time Supported' value {data['Time Supported']}")
        self.parameter_table = False
        self.commands = False
        self.files = False
        if data["Application Identifier"]:
            self.has_application_identifier = True
            self.application_identifier = data["Application Identifier"]
        if data["Endpoints"]:
            self.endpoints = data["Endpoints"]

    def add_supported_service(self, service):
        match service:
            case "Parameter Table":
                self.parameter_table = True
            case "Files":
                self.files = True
            case "Commands":
                self.commands = True
            case _:
                raise ValueError(f"Unexpected service {service}")

    def as_c_reach_struct(self, name, indent: int):
        fields = self.__dict__
        # String fields need to be modified to be displayed correctly
        for key in fields.keys():
            match type(fields[key]):
                case builtins.str:
                    # Surround with quotes
                    fields[key] = f'"{fields[key]}"'
                case builtins.bool:
                    if fields[key]:
                        fields[key] = "true"
                    else:
                        fields[key] = "false"
        fields['protocol_version'] = "cr_ReachProtoVersion_CURRENT_VERSION"
        services = []
        if fields['parameter_table'] == "true":
            services.append("cr_ServiceIds_PARAMETER_REPO")
        if fields['files'] == "true":
            services.append("cr_ServiceIds_FILES")
        if fields['commands'] == "true":
            services.append("cr_ServiceIds_COMMANDS")
        if fields['cli'] == "true":
            services.append("cr_ServiceIds_CLI")
        if fields['time'] == "true":
            services.append("cr_ServiceIds_TIME")
        del fields['parameter_table']
        del fields['files']
        del fields['commands']
        del fields['cli']
        del fields['time']
        fields['services'] = " | ".join(services)
        lines = []
        max_len = len(max(fields.keys(), key=len))
        for key in fields.keys():
            lines.append(f"{INDENT * (indent + 1)}.{key.ljust(max_len)} = {fields[key]}")
        separator = ",\n"
        header = f"{INDENT * indent}const cr_DeviceInfoResponse {name} = \n{INDENT * indent}{{"
        footer = f"{INDENT * indent}}};"
        return f"{header}\n{separator.join(lines)}\n{footer}"


class CrParamExInfo:
    def __init__(self, name: str, data):
        # Temporarily use the data to kill a warning about this being unused
        self.name = data
        self.name = name
        self.params = []
        self.names = []
        self.enums = []
        self.values = []
        self.num_structs = 0
        self.data_type = None

    def add_param(self, param: 'CrParameterInfo'):
        if param in self.params:
            raise ValueError(f"Parameter ID '{param.id}' is already in this list")
        self.params.append(param)

    def as_c_enum(self, indent: int):
        return gen_enum(self.enums, self.values, self.name, indent=indent)

    def as_c_reach_struct(self, indent: int):
        max_enum_length = len(max(self.enums, key=len))
        lines = []
        for param in self.params:
            start_index = 0
            if len(self.enums) <= 7:
                end_index = len(self.enums) - 1
            else:
                end_index = 7
            finished = False
            while not finished:
                if end_index == len(self.names) - 1:
                    # Break out of loop after this iteration, out of definitions
                    finished = True
                lines.append(f"{INDENT * indent}{{")
                lines.append(f"{INDENT * (indent + 1)}.associated_pid = {param.id},")
                lines.append(f"{INDENT * (indent + 1)}.data_type = {self.data_type},")
                lines.append(f"{INDENT * (indent + 1)}.enumerations_count = {end_index - start_index + 1},")
                lines.append(f"{INDENT * (indent + 1)}.enumerations = {{")
                for i in range(end_index - start_index + 1):
                    index = start_index + i
                    if i == end_index - start_index:
                        end = ""
                    else:
                        end = ","
                    temp = f"{self.enums[index]},"
                    lines.append(f'{INDENT * (indent + 2)}{{'
                                 f'{temp.ljust(max_enum_length + 1)} "{self.names[index]}"}}{end}')
                lines.append(f"{INDENT * (indent + 1)}}}")
                lines.append(f"{INDENT * indent}}},")
                start_index += 8
                if len(self.names) <= start_index + 8:
                    end_index = len(self.names) - 1
                else:
                    end_index += 8
        separator = "\n"
        output = separator.join(lines)
        output = output[:-1]
        return output


class CrEnumInfo(CrParamExInfo):
    def __init__(self, name: str, data):
        super().__init__(name, data)
        self.data_type = "cr_ParameterDataType_ENUMERATION"
        current_index = 0
        for row in data:
            self.names.append(row['Name'])
            if current_index % 8 == 0:
                self.num_structs += 1
            if row['Value']:
                if row['Value'] <= current_index:
                    raise ValueError(f"Enumeration values must be increasing.  "
                                     f"The last value was {current_index}, the new value is {row['Value']}")
                self.values.append(row['Value'])
                current_index = row['Value']
            else:
                current_index += 1
                self.values.append(None)
            temp_enum = f"{self.name}_{row['Name']}"
            self.enums.append(make_c_compatible(temp_enum, upper=True))

    def __repr__(self):
        return f"'{self.name}' enumeration with {len(self.names)} elements: {self.names}"


class CrBitfieldInfo(CrParamExInfo):
    def __init__(self, name: str, data):
        super().__init__(name, data)
        self.data_type = "cr_ParameterDataType_BIT_FIELD"
        current_index = 0
        for row in data:
            self.names.append(row['Name'])
            if current_index % 8 == 0:
                self.num_structs += 1
            if row['Bit Index']:
                if row['Bit Index'] < current_index:
                    raise ValueError(f"Bitfield indices must be increasing.  "
                                     f"The last index was {current_index}, the new index is {row['Value']}")
                current_index = row['Bit Index']
            self.values.append(f"(0x1 << {int(current_index)})")
            temp_enum = f"{self.name}_BIT_{row['Name']}"
            self.enums.append(make_c_compatible(temp_enum, upper=True))
            current_index += 1

    def __repr__(self):
        return f"'{self.name}' bitfield with {len(self.names)} elements: {self.names}"


class CrParameterInfo:
    def __init__(self, row, ext_types):
        if row['Name'] is None or row['Type'] is None or row['Access'] is None or row['Storage Location'] is None:
            raise ValueError(f"Input data is missing required values.  Input: {row}")
        self.name = row['Name']
        temp_id = f"PARAM_{self.name}"
        self.id = make_c_compatible(temp_id, upper=True)
        self.type = row['Type']
        if row['Description']:
            self.description = row['Description']
        if row['Units']:
            self.units = row['Units'].replace("°", r"\xC2\xB0")
        if row['Size']:
            self.size = int(row['Size'])
        self.access = row['Access']
        self.location = row['Storage Location']
        if row['Minimum Value'] is not None:
            self.min_value = row['Minimum Value']
        if row['Default Value'] is not None:
            self.default_value = row['Default Value']
        if row['Maximum Value'] is not None:
            self.max_value = row['Maximum Value']

        match self.type:
            case x if x in ["enum", "bitfield"]:
                if row['Extended Type']:
                    self.ext_type = row['Extended Type']
                else:
                    self.ext_type = row["Name"]
                definition_found = False
                for ext_type in ext_types:
                    if ext_type.name == self.ext_type:
                        if (self.type == "enum" and type(ext_type) is CrEnumInfo) or \
                                (self.type == "bitfield" and type(ext_type) is CrBitfieldInfo):
                            ext_type.add_param(self)
                            definition_found = True
                            break
                        else:
                            raise ValueError(f"Extended type mismatch for parameter "
                                             f"'{self.name}' with type '{self.ext_type}'")
                if not definition_found:
                    self.ext_type = None
            case x if x in ["uint32", "int32", "float32", "uint64", "int64", "bool", "string"]:
                # No additional validation required
                pass
            case "bytes":
                if self.size is None:
                    raise ValueError("Bytes type given with no provided size")

    def __repr__(self):
        return f"{self.name} ({self.type})"

    def as_c_reach_param_info(self, indent: int):
        struct_fields = [
            "id",
            "data_type",
            "size_in_bytes",
            "name",
            "access",
            "description",
            "units",
            "has_description",
            "has_range_min",
            "has_range_max",
            "has_default_value",
            "range_min",
            "range_max",
            "storage_location"
        ]

        data_types = {
            "uint32": "cr_ParameterDataType_UINT32",
            "int32": "cr_ParameterDataType_INT32",
            "float32": "cr_ParameterDataType_FLOAT32",
            "uint64": "cr_ParameterDataType_UINT64",
            "int64": "cr_ParameterDataType_INT64",
            "float64": "cr_ParameterDataType_FLOAT64",
            "bool": "cr_ParameterDataType_BOOL",
            "string": "cr_ParameterDataType_STRING",
            "enum": "cr_ParameterDataType_ENUMERATION",
            "bitfield": "cr_ParameterDataType_BIT_FIELD",
            "bytes": "cr_ParameterDataType_BYTE_ARRAY"
        }

        access_levels = {
            "None": "cr_AccessLevel_NO_ACCESS",
            "Read": "cr_AccessLevel_READ",
            "Write": "cr_AccessLevel_WRITE",
            "Read/Write": "cr_AccessLevel_READ_WRITE"
        }

        storage_locations = {
            "RAM": "cr_StorageLocation_RAM",
            "Extended RAM": "cr_StorageLocation_RAM_EXTENDED",
            "NVM": "cr_StorageLocation_NONVOLATILE",
            "Extended NVM": "cr_StorageLocation_NONVOLATILE_EXTENDED"
        }

        lines = []
        max_len = len(max(struct_fields, key=len))
        lines.append(f"{INDENT * (indent + 1)}.{'id'.ljust(max_len)} = {self.id},")
        lines.append(f"{INDENT * (indent + 1)}.{'name'.ljust(max_len)} = \"{self.name}\",")
        lines.append(f"{INDENT * (indent + 1)}.{'data_type'.ljust(max_len)} = {data_types[self.type]},")
        if (self.type == "enum" or self.type == "bitfield") and self.ext_type is None:
            lines[-1] += f" // {self.type} not documented"
        if 'size' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'size_in_bytes'.ljust(max_len)} = {self.size},")
        lines.append(f"{INDENT * (indent + 1)}.{'access'.ljust(max_len)} = {access_levels[self.access]},")
        lines.append(f"{INDENT * (indent + 1)}"
                     f".{'storage_location'.ljust(max_len)} = {storage_locations[self.location]},")
        if 'description' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'has_description'.ljust(max_len)} = true,")
            lines.append(f'{INDENT * (indent + 1)}.{"description".ljust(max_len)} = "{self.description}",')
        if 'units' in self.__dict__:
            lines.append(f'{INDENT * (indent + 1)}.{"units".ljust(max_len)} = "{self.units}",')
        if 'min_value' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'has_range_min'.ljust(max_len)} = true,")
            lines.append(f'{INDENT * (indent + 1)}.{"range_min".ljust(max_len)} = {self.min_value},')
        if 'default_value' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'has_default_value'.ljust(max_len)} = true,")
            lines.append(f'{INDENT * (indent + 1)}.{"default_value".ljust(max_len)} = {self.default_value},')
        if 'max_value' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'has_range_max'.ljust(max_len)} = true,")
            lines.append(f'{INDENT * (indent + 1)}.{"range_max".ljust(max_len)} = {self.max_value},')
        separator = "\n"
        header = f"{INDENT * indent}{{"
        footer = f"{INDENT * indent}}}"
        return f"{header}\n{separator.join(lines)}\n{footer}"


class CrFile:
    def __init__(self, data: dict):
        self.name = data["Name"]
        temp_id = f"FILE_{self.name}"
        self.id = make_c_compatible(temp_id, upper=True)
        self.access = data["Access"]
        self.maxSize = data["Max Size"]
        self.size = int(data["Size"])
        self.location = data["Storage Location"]
        self.require_checksum = "false"
        if data["Require Checksum"] == "On":
            self.require_checksum = "true"

    def as_c_info(self, indent: int):
        struct_fields = [
            "file_id",
            "file_name",
            "access",
            "current_size_bytes",
            "maximum_size_bytes",
            "storage_location",
            "require_checksum"
        ]

        access_levels = {
            "None": "cr_AccessLevel_NO_ACCESS",
            "Read": "cr_AccessLevel_READ",
            "Write": "cr_AccessLevel_WRITE",
            "Read/Write": "cr_AccessLevel_READ_WRITE"
        }

        storage_locations = {
            "RAM": "cr_StorageLocation_RAM",
            "Extended RAM": "cr_StorageLocation_RAM_EXTENDED",
            "NVM": "cr_StorageLocation_NONVOLATILE",
            "Extended NVM": "cr_StorageLocation_NONVOLATILE_EXTENDED"
        }

        lines = []
        max_len = len(max(struct_fields, key=len))
        lines.append(f"{INDENT * (indent + 1)}.{'file_id'.ljust(max_len)} = {self.id},")
        lines.append(f"{INDENT * (indent + 1)}.{'file_name'.ljust(max_len)} = \"{self.name}\",")
        if 'size' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'current_size_bytes'.ljust(max_len)} = {self.size},")
        if 'maxSize' in self.__dict__:
            lines.append(f"{INDENT * (indent + 1)}.{'maximum_size_bytes'.ljust(max_len)} = {self.maxSize},")
        lines.append(f"{INDENT * (indent + 1)}.{'access'.ljust(max_len)} = {access_levels[self.access]},")
        lines.append(f"{INDENT * (indent + 1)}"
                     f".{'storage_location'.ljust(max_len)} = {storage_locations[self.location]},")
        lines.append(f"{INDENT * (indent + 1)}"
                     f".{'require_checksum'.ljust(max_len)} = {self.require_checksum},")
        separator = "\n"
        header = f"{INDENT * indent}{{"
        footer = f"{INDENT * indent}}}"
        return f"{header}\n{separator.join(lines)}\n{footer}"


class CrCommand:
    def __init__(self, data: dict):
        self.name = data["Name"]
        temp_id = f"COMMAND_{self.name}"
        self.id = make_c_compatible(temp_id, upper=True)

    def as_c_info(self, indent: int):
        struct_fields = [
            "id",
            "name",
        ]

        lines = []
        max_len = len(max(struct_fields, key=len))
        lines.append(f"{INDENT * (indent + 1)}.{'id'.ljust(max_len)} = {self.id},")
        lines.append(f"{INDENT * (indent + 1)}.{'name'.ljust(max_len)} = \"{self.name}\",")
        separator = "\n"
        header = f"{INDENT * indent}{{"
        footer = f"{INDENT * indent}}}"
        return f"{header}\n{separator.join(lines)}\n{footer}"


class ReachDevice:
    def __init__(self, filename):
        def parse_sheet(sheet, column_based=True, expected_header=None):
            if column_based:
                header = [
                    c
                    for c in (cell.value for cell in sheet[1])
                    if c not in [None]
                ]
                if expected_header:
                    if header != expected_header:
                        raise ValueError("Unexpected header contents")
                post_process_funcs = [lambda x: x] * len(header)
                dicts = []
                for row_idx in range(2, sheet.max_row + 1):
                    candidate = {}
                    for col_idx in range(1, sheet.max_column + 1):
                        if col_idx <= len(header):
                            candidate[header[col_idx - 1]] = post_process_funcs[
                                col_idx - 1
                                ](sheet.cell(row=row_idx, column=col_idx).value)
                    if not all(value in [None] for value in candidate.values()):
                        dicts.append(candidate)
                return dicts
            else:
                result = {}
                if sheet.max_row != len(expected_header):
                    raise ValueError("Unexpected sheet contents")
                for row, hdr in zip(sheet, expected_header):
                    if row[0].value != hdr:
                        raise ValueError(f"Expected {hdr} but found {row[0].value}")
                    result[hdr] = row[1].value
                return result

        book = load_workbook(filename, data_only=True)
        sheets = book.sheetnames
        self.ext_types = []
        self.params = []
        self.device_info = None
        self.files = None
        self.commands = None
        params_dicts = None
        for s in sheets:
            match s:
                case "Device Information":
                    self.device_info = CrDeviceInfo(
                        parse_sheet(book[s], column_based=False,
                                    expected_header=["Device Name", "Manufacturer", "Device Description",
                                                     "CLI Supported", "Time Supported", "Application Identifier",
                                                     "Endpoints"]))
                case "Parameters":
                    params_dicts = parse_sheet(book[s],
                                               expected_header=["Name", "Type", "Extended Type", "Description",
                                                                "Units", "Size", "Access", "Minimum Value",
                                                                "Default Value", "Maximum Value", "Storage Location",
                                                                "Notify", "min period", "max period", "delta"])
                case "Files":
                    files = parse_sheet(book[s], expected_header=["Name", "Size", "Access", "Storage Location", 
                                                                  "Require Checksum", "Max Size"])
                    self.files = []
                    for file in files:
                        self.files.append(CrFile(file))
                case "Commands":
                    temp = parse_sheet(book[s], expected_header=["Name", "Description"])
                    self.commands = []
                    for cmd in temp:
                        self.commands.append(CrCommand(cmd))
                case s if s in ["Example Enumeration", "Example Bitfield"]:
                    # Example tabs, ignore
                    continue
                case _:
                    # Try parsing this as an enum and then as a bitfield if that doesn't work
                    try:
                        self.ext_types.append(CrEnumInfo(s, parse_sheet(book[s], expected_header=["Name", "Value"])))
                    except ValueError:
                        self.ext_types.append(CrBitfieldInfo(s, parse_sheet(book[s],
                                                                            expected_header=["Name", "Bit Index"])))

        if not self.device_info:
            raise ValueError("No device information was provided")
        for param in params_dicts:
            self.params.append(CrParameterInfo(param, self.ext_types))
        if params_dicts:
            self.device_info.add_supported_service("Parameter Table")
        if self.files:
            self.device_info.add_supported_service("Files")
        if self.commands:
            self.device_info.add_supported_service("Commands")

    def create_files(self, h_location="./", c_location="./"):
        header_text = "A minimal implementation of Reach data access.  Auto-generated by a Python script."

        def c_file():
            includes = '#include "definitions.h"\n#include <stdio.h>\n#include <string.h>\n#include <assert.h>\n' \
                       '#include "i3_log.h"\n#include "app_version.h"\n#include "cr_stack.h"'
            define_names = []
            define_values = []
            definitions = [f"{self.device_info.as_c_reach_struct('device_info', 0)}"]
            main_functions = [f"{ccgu.DeviceInfo.main_functions}"]
            weak_functions = []
            if self.params:
                param_lines = []
                for param in self.params:
                    param_lines.append(param.as_c_reach_param_info(1))
                separator = ",\n"
                definitions.append(f"cr_ParameterValue sCr_param_val[NUM_PARAMS];")
                definitions.append(
                    f"cr_ParameterInfo param_desc[NUM_PARAMS] = {{\n{separator.join(param_lines)}\n}};")
                main_functions.append(ccgu.ParamRepo.main_functions)
                weak_functions.append(ccgu.ParamRepo.weak_access_functions)
                if self.ext_types:
                    enum_struct_count = 0
                    for ext_type in self.ext_types:
                        enum_struct_count += ext_type.num_structs * len(ext_type.params)
                    ext_lines = []
                    for ext in self.ext_types:
                        temp = ext.as_c_reach_struct(1)
                        if temp == "":
                            # Unused extended type
                            continue
                        ext_lines.append(temp)
                    definitions.append(f"cr_ParamExInfoResponse param_ex_desc[NUM_EX_PARAMS] = {{\n"
                                       f"{separator.join(ext_lines)}\n}};")
                    main_functions.append(ccgu.ParamRepo.ext_param_functions)
            if self.files:
                file_lines = []
                for file in self.files:
                    file_lines.append(file.as_c_info(1))
                separator = ",\n"
                definitions.append(f"cr_FileInfo file_descriptions[NUM_FILES] = {{\n{separator.join(file_lines)}\n}};")
                main_functions.append(ccgu.Files.main_functions)
            if self.commands:
                cmd_lines = []
                for cmd in self.commands:
                    cmd_lines.append(cmd.as_c_info(1))
                separator = ",\n"
                definitions.append(f"cr_CommandInfo command_desc[NUM_COMMANDS] = {{\n"
                                   f"{separator.join(cmd_lines)}\n}};")
                main_functions.append(ccgu.Commands.main_functions)

            if len(define_names) > 0:
                defines = []
                max_define_len = len(max(define_names, key=len))
                for name, value in zip(define_names, define_values):
                    defines.append(f"#define {name.ljust(max_define_len)} {value}")
                defines = "\n".join(defines)
                defines += "\n\n"
            else:
                defines = ""

            if len(definitions) > 0:
                definitions = "\n\n".join(definitions)
                definitions += "\n\n"
            else:
                definitions = ""

            if len(main_functions) > 0:
                main_functions = "\n\n".join(main_functions)
                main_functions += "\n\n"
            else:
                main_functions = ""

            if len(weak_functions) > 0:
                weak_functions = "\n\n".join(weak_functions)
            else:
                weak_functions = ""

            header = gen_header("definitions.c", header_text)
            output = f"{header}\n\n{includes}\n\n{defines}{definitions}{main_functions}{weak_functions}"
            return output

        def h_file():
            start_guard = "#ifndef _DEFINITIONS_H\n#define _DEFINITIONS_H"
            end_guard = "#endif // _DEFINITIONS_H"
            includes = r'''#include "reach.pb.h"'''

            define_names = []
            define_values = []
            variables = []
            enums = []
            access_functions = []
            if self.params:
                define_names.append("INCLUDE_PARAMETER_SERVICE")
                define_values.append(None)
                define_names.append(f"NUM_PARAMS")
                define_values.append(f"{len(self.params)}")
                enum_names = []
                enum_values = []
                for param in self.params:
                    enum_names.append(param.name)
                    enum_values.append(None)
                enums.append(gen_enum(enum_names, enum_values, "param", transform_enum_names=True))
                access_functions.append(ccgu.ParamRepo.weak_access_functions_h)
                variables.append(f"extern cr_ParameterValue sCr_param_val[NUM_PARAMS];")
                variables.append(f"extern cr_ParameterInfo param_desc[NUM_PARAMS];")
                if self.ext_types:
                    enum_struct_count = 0
                    for ext_type in self.ext_types:
                        enum_struct_count += ext_type.num_structs * len(ext_type.params)
                    define_names.append("NUM_EX_PARAMS")
                    define_values.append(f"{enum_struct_count}")
                    for ext_type in self.ext_types:
                        enums.append(ext_type.as_c_enum(0))
                    variables.append(f"extern cr_ParamExInfoResponse param_ex_desc[NUM_EX_PARAMS];")
            if self.files:
                define_names.append("INCLUDE_FILE_SERVICE")
                define_values.append(None)
                define_names.append("NUM_FILES")
                define_values.append(f"{len(self.files)}")
                enum_names = []
                enum_values = []
                for file in self.files:
                    enum_names.append(file.name)
                    enum_values.append(None)
                enums.append(gen_enum(enum_names, enum_values, "file", transform_enum_names=True))
                variables.append(f"extern cr_FileInfo file_descriptions[NUM_FILES];")
            if self.commands:
                define_names.append("INCLUDE_COMMAND_SERVICE")
                define_values.append(None)
                define_names.append("NUM_COMMANDS")
                define_values.append(f"{len(self.commands)}")
                enum_names = []
                enum_values = []
                for cmd in self.commands:
                    enum_names.append(cmd.id)
                    enum_values.append(None)
                enums.append(gen_enum(enum_names, enum_values, "command"))
                variables.append(f"extern cr_CommandInfo command_desc[NUM_COMMANDS];")
            if self.device_info.cli:
                define_names.append("INCLUDE_CLI_SERVICE")
                define_values.append(None)
            if self.device_info.time:
                define_names.append("INCLUDE_TIME_SERVICE")
                define_values.append(None)
            if len(define_names) > 0:
                defines = []
                max_define_len = len(max(define_names, key=len))
                for name, value in zip(define_names, define_values):
                    if value is not None:
                        defines.append(f"#define {name.ljust(max_define_len)} {value}")
                    else:
                        defines.append(f"#define {name}")
                defines = "\n".join(defines)
                defines += "\n\n"
            else:
                defines = ""

            if len(variables) > 0:
                variables = "\n".join(variables)
                variables += "\n\n"
            else:
                variables = ""

            if len(enums) > 0:
                enums = "\n\n".join(enums)
                enums += "\n\n"
            else:
                enums = ""

            if len(access_functions) > 0:
                access_functions = "\n\n".join(access_functions)
                access_functions += "\n\n"
            else:
                access_functions = ""

            header = gen_header("definitions.h", header_text)
            return f"{header}\n\n{start_guard}\n\n{includes}\n\n{defines}{enums}{variables}" \
                   f"{access_functions}{end_guard}"

        with open(f"{h_location}definitions.h", "w") as f:
            f.write(h_file())

        with open(f"{c_location}definitions.c", "w") as f:
            f.write(c_file())


parser = argparse.ArgumentParser(
                    description='A script to transform a .xlsx file defining a Reach device into C code')
parser.add_argument('-x', '--xlsx', help="The .xlsx file to parse", required=True)
parser.add_argument('-s', '--source-location', help="Where to put the generated 'definitions.c' file", default=".")
parser.add_argument('-i', '--include-location', help="Where to put the generated 'definitions.h' file", default=".")

args = parser.parse_args()

if args.source_location[-1] not in ["/", "\\"]:
    args.source_location += "/"
if args.include_location[-1] not in ["/", "\\"]:
    args.include_location += "/"

test = ReachDevice(args.xlsx)
test.create_files(h_location=args.include_location, c_location=args.source_location)
